import os
import time
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


def unfriend(x):
    # From here the whole process starts.....

    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[1]/label[1]/input[1]').click()
    time.sleep(2)
    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[1]/label[1]/input[1]').send_keys(Keys.CONTROL+"A")
    time.sleep(2)
    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[1]/label[1]/input[1]').send_keys(x)
    time.sleep(2)
    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/div[1]/div[4]/a[1]/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]').click()
    time.sleep(2)
    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[4]/div[2]/div[1]/div[1]/span[1]').click()
    time.sleep(2)
    driver.find_element(By.XPATH,'/html[1]/body[1]/div[1]/div[1]/div[1]/div[1]/div[4]/div[1]/div[1]/div[1]/div[1]/div[2]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[3]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/div[1]/span[1]/span[1]').click()
    time.sleep(2)

USER = "Enter your pc's USER NAME"

driver_path = Service('driver/chromedriver.exe')
profiles = webdriver.ChromeOptions()
profiles.add_argument("--user-data-dir=/Users/"+USER+"/AppData/Local/Google/Chrome/User Data/")
profiles.add_argument("--profile-directory=Default")
driver = webdriver.Chrome(service=driver_path, options=profiles)
link = 'https://www.facebook.com/friends/list'
driver.get(link)

time.sleep(3)
print('Page is ready to work...')

data = pd.read_excel('data_sheet.xlsx', dtype=str)

start = int(input("Enter starting Row:"))
end = int(input("Enter ending Row:"))

for row in range(start - 1, end):
    name = data['Name'][row]
    status = 'Unfriend'

    unfriend(name)

    # Storing Unfriend data into Unfriend_status excel file.....

    if os.path.exists('unfriend_status.xlsx'):

        wb = openpyxl.load_workbook('unfriend_status.xlsx')
        ws = wb.active
        data_list = [[name, status]]
        for dat in data_list:
            ws.append(dat)
        wb.save('unfriend_status.xlsx')


    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['B1'] = 'Status'
        wb = openpyxl.load_workbook('unfriend_status.xlsx')
        ws = wb.active
        data_list = [[name, status]]
        for dat in data_list:
            ws.append(dat)
        wb.save('unfriend_status.xlsx')

    print(name)

driver.close()
print('Given Work is done.....')





